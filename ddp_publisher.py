#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
DDP PUBLISHER - pubblicazione automatica dei Social Pack "Diari di Pelle"
su Instagram, Facebook e TikTok.

Legge il calendario Excel gia' pronto (Calendario_Social_Diari_di_Pelle_4_Mesi.xlsx),
lo trasforma in una coda di lavoro, e pubblica ogni contenuto alla data/ora prevista.

Comandi
-------
  python ddp_publisher.py build   --config config.json
  python ddp_publisher.py check   --config config.json
  python ddp_publisher.py plan    --config config.json [--days 7]
  python ddp_publisher.py run     --config config.json [--dry-run] [--window 30]
  python ddp_publisher.py fb-batch --config config.json [--days 30] [--dry-run]
  python ddp_publisher.py status  --config config.json

Filosofia: niente magie. Ogni azione e' tracciata in state.json, niente viene
pubblicato due volte, e --dry-run stampa esattamente cosa verrebbe inviato.
"""

from __future__ import annotations

import argparse
import json
import mimetypes
import os
import re
import sys
import time
from dataclasses import dataclass, asdict, field
from datetime import datetime, timedelta, date, time as dtime
from pathlib import Path
from typing import Any, Dict, List, Optional
from urllib.parse import urlencode, unquote

try:
    import requests
except ImportError:
    print("Manca la libreria requests:  py -m pip install requests")
    raise

try:
    import openpyxl
except ImportError:
    openpyxl = None


GRAPH = "https://graph.facebook.com/v21.0"
TIKTOK = "https://open.tiktokapis.com/v2"


def _pulisci_token(grezzo: str) -> str:
    righe = [r.strip() for r in grezzo.splitlines()
             if r.strip() and not r.strip().startswith("#")]
    return "".join(righe).replace(" ", "").replace("\t", "")


def leggi_token(cfg, piattaforma: str) -> str:
    """
    Prende il token da un file esterno se c'e', altrimenti da config.json.
    Il file ha la precedenza: cosi' non devi mai incollare token dentro il JSON.

    File cercati (nella cartella del config):
      facebook  -> token_facebook.txt
      instagram -> token_instagram.txt   (se assente usa quello di facebook)
      tiktok    -> token_tiktok.txt
    """
    sezione = cfg.get(piattaforma, {})

    # 1) variabile d'ambiente (usata dai "segreti" di GitHub)
    da_ambiente = os.environ.get(f"TOKEN_{piattaforma.upper()}", "").strip()
    if da_ambiente:
        return _pulisci_token(da_ambiente)

    # 2) file di testo nella cartella
    nome = sezione.get("token_file") or f"token_{piattaforma}.txt"
    f = Path(cfg["_dir"]) / nome
    if f.exists():
        grezzo = f.read_text(encoding="utf-8")
        if _pulisci_token(grezzo):
            return _pulisci_token(grezzo)

    chiave = "page_access_token" if piattaforma == "facebook" else "access_token"
    return (sezione.get(chiave) or "").strip()

# ---------------------------------------------------------------------------
# Modello dati
# ---------------------------------------------------------------------------


@dataclass
class Task:
    task_id: str
    when: str                 # ISO "2026-07-28T20:00:00"
    platform: str             # instagram | facebook | tiktok
    fmt: str                  # reel | storia | carosello
    book: str
    content_id: str           # TRAILER, CH01, ...
    title: str
    caption: str = ""
    link: str = ""
    media: List[str] = field(default_factory=list)   # percorsi relativi alla root dei pack
    cover: Optional[str] = None

    @property
    def dt(self) -> datetime:
        return datetime.fromisoformat(self.when)

    @property
    def key(self) -> str:
        return f"{self.task_id}|{self.platform}|{self.fmt}|{self.content_id}"


# ---------------------------------------------------------------------------
# Config e stato
# ---------------------------------------------------------------------------


def load_config(path: Path) -> Dict[str, Any]:
    cfg = json.loads(path.read_text(encoding="utf-8"))
    cfg["_dir"] = str(path.parent.resolve())
    return cfg


def _state_path(cfg) -> Path:
    return Path(cfg["_dir"]) / cfg.get("state_file", "state.json")


def load_state(cfg) -> Dict[str, Any]:
    p = _state_path(cfg)
    if p.exists():
        return json.loads(p.read_text(encoding="utf-8"))
    return {"done": {}, "failed": {}}


def save_state(cfg, state) -> None:
    _state_path(cfg).write_text(
        json.dumps(state, indent=2, ensure_ascii=False), encoding="utf-8"
    )


def mark(state, task: Task, ok: bool, info: str) -> None:
    bucket = "done" if ok else "failed"
    state[bucket][task.key] = {
        "at": datetime.now().isoformat(timespec="seconds"),
        "platform": task.platform,
        "format": task.fmt,
        "content": task.content_id,
        "info": info,
    }


# ---------------------------------------------------------------------------
# BUILD: dal calendario Excel alla coda JSON
# ---------------------------------------------------------------------------

FORMAT_MAP = {
    "reel": "reel",
    "video": "reel",
    "storia": "storia",
    "story": "storia",
    "carosello": "carosello",
    "carousel": "carosello",
}

PLATFORM_MAP = {
    "instagram": "instagram",
    "facebook": "facebook",
    "tiktok": "tiktok",
}


def _cell_text(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.isoformat()
    return str(v).strip()


def _resolve(value, sheets: Dict[str, Any], depth: int = 0):
    """
    Segue i riferimenti tipo ='TESTI'!E2 o ='CAMPAGNE'!H2 e restituisce il valore
    vero della cella puntata. Ricorsivo: se anche quella e' un riferimento, tira
    avanti (fino a 5 salti, per non girare in tondo).
    Se non e' un riferimento, restituisce il valore com'e'.
    """
    if depth > 5 or not isinstance(value, str) or not value.startswith("="):
        return value

    testo = value.strip()

    # ='ALTRO FOGLIO'!E2   oppure   =ALTROFOGLIO!E2
    m = re.match(r"^='?([^'!]+)'?!\$?([A-Z]+)\$?(\d+)$", testo)
    if m:
        ws = sheets.get(m.group(1))
        if ws is None:
            return None
        return _resolve(ws[f"{m.group(2)}{m.group(3)}"].value, sheets, depth + 1)

    return None


def _resolve_text(value, sheets: Dict[str, Any]) -> str:
    return _cell_text(_resolve(value, sheets))


def _as_date(value) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str) and value.strip():
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(value.strip()[:10], fmt).date()
            except ValueError:
                continue
    return None


def _as_time(value) -> Optional[dtime]:
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, dtime):
        return value
    if isinstance(value, str) and ":" in value:
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                return datetime.strptime(value.strip(), fmt).time()
            except ValueError:
                continue
    return None


def build_queue(cfg) -> List[Task]:
    if openpyxl is None:
        sys.exit("Manca openpyxl:  py -m pip install openpyxl")

    xlsx = Path(cfg["calendar_xlsx"])
    if not xlsx.is_absolute():
        xlsx = Path(cfg["_dir"]) / xlsx

    # Due letture dello stesso file:
    #  - wb_val: i RISULTATI che Excel ha gia' calcolato e salvato (le date, le caption)
    #  - wb_fml: le formule, usate come riserva se un risultato non e' stato salvato
    wb_val = openpyxl.load_workbook(xlsx, data_only=True)
    wb_fml = openpyxl.load_workbook(xlsx, data_only=False)
    sheets = {ws.title: ws for ws in wb_fml.worksheets}

    cal = wb_val["CALENDARIO"] if "CALENDARIO" in wb_val.sheetnames else None
    if cal is None:
        sys.exit("Nel file Excel manca il foglio CALENDARIO.")
    cal_fml = sheets["CALENDARIO"]

    headers = [_cell_text(c.value) for c in cal_fml[1]]
    idx = {h: i for i, h in enumerate(headers)}

    def col(*nomi):
        """Prende la prima colonna che esiste tra i nomi indicati.
        Serve perche' esistono due versioni del calendario, con nomi diversi."""
        for n in nomi:
            if n in idx:
                return idx[n]
        return None

    C = {
        "task":     col("Task"),
        "data":     col("Data pubblicazione", "Data"),
        "ora":      col("Ora pubblicazione", "Ora"),
        "piatt":    col("Piattaforma"),
        "formato":  col("Formato"),
        "libro":    col("Libro"),
        "contenuto": col("Contenuto"),
        "titolo":   col("Titolo"),
        "file":     col("File principale"),
        "extra":    col("Copertina / file aggiuntivo"),
        "caption":  col("Caption / testo da copiare", "Caption"),
        "link":     col("Link"),
        "stato":    col("Stato"),
    }

    mancanti = [k for k, v in C.items()
                if v is None and k in ("task", "data", "piatt", "formato", "file")]
    if mancanti:
        print("ATTENZIONE: nel CALENDARIO non trovo queste colonne: "
              + ", ".join(mancanti))
        print("Colonne presenti: " + ", ".join(h for h in headers if h))
        sys.exit(1)

    packs_root = Path(cfg["packs_root"])
    tasks: List[Task] = []
    scartate = 0
    senza_valore = 0

    for row in cal.iter_rows(min_row=2):
        riga = row[0].row
        cells = [c.value for c in row]

        def get(chiave):
            """Prende il risultato calcolato. Se Excel non l'ha salvato,
            ripiega sulla formula e prova a seguirla."""
            nonlocal senza_valore
            i = C.get(chiave)
            if i is None or i >= len(cells):
                return None
            v = cells[i]
            if v is None:
                grezzo = cal_fml.cell(riga, i + 1).value
                if isinstance(grezzo, str) and grezzo.startswith("="):
                    senza_valore += 1
                    return _resolve(grezzo, sheets)
                return grezzo
            return v

        task_id = _cell_text(get("task"))
        if not task_id:
            continue

        platform = PLATFORM_MAP.get(_cell_text(get("piatt")).lower())
        fmt = FORMAT_MAP.get(_cell_text(get("formato")).lower())
        if not platform or not fmt:
            scartate += 1
            continue

        day = _as_date(get("data"))
        if day is None:
            scartate += 1
            continue
        clock = _as_time(get("ora")) or dtime(20, 0)
        when = datetime.combine(day, clock)

        content_id = _cell_text(get("contenuto"))
        main_file = _cell_text(get("file"))
        extra_file = _cell_text(get("extra"))

        caption = _cell_text(get("caption"))
        link = _cell_text(get("link"))

        # La cella "File principale" puo' contenere piu' percorsi su righe separate
        # (e' il caso dei caroselli: 01, 02, 03).
        listed = [x.strip() for x in re.split(r"[\r\n]+", main_file) if x.strip()]

        media: List[str] = []
        cover: Optional[str] = None
        if fmt == "carosello":
            if len(listed) > 1:
                media = listed
            elif listed:
                folder = (packs_root / listed[0]).parent
                if folder.is_dir():
                    media = [
                        str(p.relative_to(packs_root)).replace("\\", "/")
                        for p in sorted(folder.glob("*.jpg"))
                    ]
                else:
                    media = listed
        else:
            media = listed[:1]
            if extra_file and ("/" in extra_file or "\\" in extra_file):
                if extra_file.lower().endswith((".jpg", ".jpeg", ".png")):
                    cover = extra_file

        tasks.append(
            Task(
                task_id=task_id,
                when=when.isoformat(timespec="seconds"),
                platform=platform,
                fmt=fmt,
                book=_cell_text(get("libro")),
                content_id=content_id,
                title=_cell_text(get("titolo")),
                caption=caption,
                link=link,
                media=media,
                cover=cover,
            )
        )

    tasks.sort(key=lambda x: (x.dt, x.platform))
    out = Path(cfg["_dir"]) / cfg.get("queue_file", "queue.json")
    out.write_text(
        json.dumps([asdict(t) for t in tasks], indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    print(f"Coda generata: {out}  ({len(tasks)} task)")
    if scartate:
        print(f"  ({scartate} righe saltate: piattaforma, formato o data non leggibili)")
    if not tasks:
        print("\nNESSUN TASK TROVATO. Lancia 'py diagnosi.py' e controlla che nel")
        print("CALENDARIO ci siano le colonne Piattaforma, Formato e la data.")
    return tasks


def load_queue(cfg) -> List[Task]:
    p = Path(cfg["_dir"]) / cfg.get("queue_file", "queue.json")
    if not p.exists():
        return build_queue(cfg)
    raw = json.loads(p.read_text(encoding="utf-8"))
    return [Task(**r) for r in raw]


# ---------------------------------------------------------------------------
# CHECK: validazione prima di toccare le API
# ---------------------------------------------------------------------------


def media_url(cfg, rel: str) -> str:
    """URL pubblico del file. Instagram e TikTok devono poterlo scaricare da internet."""
    base = cfg["media_base_url"].rstrip("/")
    return f"{base}/{rel.lstrip('/')}"


def cmd_check(cfg) -> int:
    tasks = load_queue(cfg)
    packs_root = Path(cfg["packs_root"])
    problems = 0

    for t in tasks:
        for rel in t.media:
            if not (packs_root / rel).exists():
                print(f"[FILE MANCANTE] {t.task_id} {t.platform}/{t.fmt} -> {rel}")
                problems += 1
        if t.cover and not (packs_root / t.cover).exists():
            print(f"[COVER MANCANTE] {t.task_id} -> {t.cover}")
            problems += 1
        if not t.caption and t.fmt != "storia":
            print(f"[CAPTION VUOTA] {t.task_id} {t.platform}/{t.content_id}")
            problems += 1
        if t.platform == "instagram" and len(t.caption) > 2200:
            print(f"[CAPTION TROPPO LUNGA IG] {t.task_id} ({len(t.caption)} caratteri)")
            problems += 1
        if t.platform == "tiktok" and len(t.caption) > 2200:
            print(f"[CAPTION TROPPO LUNGA TIKTOK] {t.task_id}")
            problems += 1
        if t.fmt == "carosello" and not (2 <= len(t.media) <= 10):
            print(f"[CAROSELLO ANOMALO] {t.task_id}: {len(t.media)} immagini")
            problems += 1

    print(f"\n{len(tasks)} task controllati, {problems} problemi.")

    # Verifica che i media siano davvero raggiungibili dall'esterno.
    if cfg.get("media_base_url") and tasks:
        sample = media_url(cfg, tasks[0].media[0])
        try:
            r = requests.head(sample, timeout=15, allow_redirects=True)
            ok = r.status_code == 200
            print(f"Media pubblici: {'OK' if ok else 'HTTP ' + str(r.status_code)}  ({sample})")
            if not ok:
                problems += 1
        except Exception as e:
            print(f"Media pubblici NON raggiungibili: {e}")
            problems += 1

    return 1 if problems else 0


def cmd_plan(cfg, args) -> None:
    tasks = filter_tasks(load_queue(cfg), args)
    state = load_state(cfg)
    now = datetime.now()
    limit = now + timedelta(days=args.days)
    print(f"Prossimi {args.days} giorni ({now:%d/%m %H:%M} -> {limit:%d/%m}):\n")
    n = 0
    for t in tasks:
        if now - timedelta(days=1) <= t.dt <= limit:
            flag = "OK " if t.key in state["done"] else ("ERR" if t.key in state["failed"] else "   ")
            print(f"{flag} {t.task_id}  {t.dt:%d/%m %H:%M}  {t.platform:9} "
                  f"{t.fmt:9} {t.content_id:8} {t.title}")
            n += 1
    if not n:
        print("(niente in questo periodo con questi filtri)")


def cmd_status(cfg) -> None:
    tasks = load_queue(cfg)
    state = load_state(cfg)
    done = sum(1 for t in tasks if t.key in state["done"])
    failed = sum(1 for t in tasks if t.key in state["failed"])
    print(f"Totale task: {len(tasks)}")
    print(f"Pubblicati:  {done}")
    print(f"Falliti:     {failed}")
    print(f"In attesa:   {len(tasks) - done - failed}")
    for k, v in state["failed"].items():
        print(f"  ERRORE {k}: {v['info'][:160]}")


# ---------------------------------------------------------------------------
# INSTAGRAM (Graph API - account Business/Creator collegato a una Pagina)
# ---------------------------------------------------------------------------


class InstagramClient:
    """
    Due modi di collegarsi:
      - via Pagina Facebook (graph.facebook.com) - serve il collegamento
      - via Instagram diretto (graph.instagram.com) - NON serve la Pagina
    Il secondo si attiva mettendo "instagram_login": true nel config.
    """

    def __init__(self, cfg):
        sez = cfg.get("instagram", {})
        self.diretto = bool(sez.get("instagram_login"))
        self.base = "https://graph.instagram.com" if self.diretto else GRAPH
        self.ig_id = "me" if self.diretto else sez.get("ig_user_id", "")
        self.token = leggi_token(cfg, "instagram") or leggi_token(cfg, "facebook")
        self.dry = False

    def _post(self, path, params):
        params = dict(params, access_token=self.token)
        if self.dry:
            safe = {k: v for k, v in params.items() if k != "access_token"}
            print(f"    POST {self.base}/{path}  {json.dumps(safe, ensure_ascii=False)[:400]}")
            return {"id": "DRYRUN"}
        r = requests.post(f"{self.base}/{path}", data=params, timeout=120)
        data = r.json()
        if "error" in data:
            raise RuntimeError(f"IG: {data['error'].get('message')}")
        return data

    def _wait_container(self, cid, timeout=600):
        if self.dry:
            return
        deadline = time.time() + timeout
        while time.time() < deadline:
            r = requests.get(
                f"{self.base}/{cid}",
                params={"fields": "status_code,status", "access_token": self.token},
                timeout=60,
            ).json()
            code = r.get("status_code")
            if code == "FINISHED":
                return
            if code == "ERROR":
                raise RuntimeError(f"IG container in errore: {r.get('status')}")
            time.sleep(6)
        raise RuntimeError("IG: timeout nell'elaborazione del video")

    def _publish(self, creation_id):
        return self._post(f"{self.ig_id}/media_publish", {"creation_id": creation_id})

    def reel(self, video_url, caption, cover_url=None):
        p = {"media_type": "REELS", "video_url": video_url, "caption": caption,
             "share_to_feed": "true"}
        if cover_url:
            p["cover_url"] = cover_url
        c = self._post(f"{self.ig_id}/media", p)
        self._wait_container(c["id"])
        return self._publish(c["id"])

    def story(self, image_url):
        c = self._post(f"{self.ig_id}/media",
                       {"media_type": "STORIES", "image_url": image_url})
        self._wait_container(c["id"])
        return self._publish(c["id"])

    def carousel(self, image_urls, caption):
        children = []
        for u in image_urls:
            c = self._post(f"{self.ig_id}/media",
                           {"image_url": u, "is_carousel_item": "true"})
            children.append(c["id"])
        parent = self._post(f"{self.ig_id}/media", {
            "media_type": "CAROUSEL",
            "children": ",".join(children),
            "caption": caption,
        })
        self._wait_container(parent["id"])
        return self._publish(parent["id"])


# ---------------------------------------------------------------------------
# FACEBOOK (Pagina - supporta la programmazione nativa)
# ---------------------------------------------------------------------------


class FacebookClient:
    def __init__(self, cfg):
        self.page_id = cfg["facebook"]["page_id"]
        self.token = leggi_token(cfg, "facebook")
        self.dry = False

    def _post(self, path, params, files=None):
        params = dict(params, access_token=self.token)
        if self.dry:
            safe = {k: v for k, v in params.items() if k != "access_token"}
            print(f"    POST {GRAPH}/{path}  {json.dumps(safe, ensure_ascii=False)[:400]}")
            return {"id": "DRYRUN"}
        r = requests.post(f"{GRAPH}/{path}", data=params, files=files, timeout=300)
        data = r.json()
        if "error" in data:
            raise RuntimeError(f"FB: {data['error'].get('message')}")
        return data

    def photo(self, image_url, message, when: Optional[datetime] = None):
        p = {"url": image_url, "caption": message}
        if when:
            p["published"] = "false"
            p["scheduled_publish_time"] = int(when.timestamp())
        return self._post(f"{self.page_id}/photos", p)

    def carousel(self, image_urls, message, when: Optional[datetime] = None):
        ids = []
        for u in image_urls:
            r = self._post(f"{self.page_id}/photos", {"url": u, "published": "false"})
            ids.append(r["id"])
        p = {"message": message}
        for i, fbid in enumerate(ids):
            p[f"attached_media[{i}]"] = json.dumps({"media_fbid": fbid})
        if when:
            p["published"] = "false"
            p["scheduled_publish_time"] = int(when.timestamp())
        return self._post(f"{self.page_id}/feed", p)

    def reel(self, video_url, description, when: Optional[datetime] = None,
             file_locale: Optional[Path] = None):
        # Reels della Pagina: init -> upload -> finish
        start = self._post(f"{self.page_id}/video_reels", {"upload_phase": "start"})
        video_id = start.get("video_id", "DRYRUN")

        if not self.dry:
            url_upload = f"https://rupload.facebook.com/video-upload/v21.0/{video_id}"

            if file_locale and file_locale.exists():
                # Mando io i byte del file. Facebook non passa dal sito, quindi
                # Hostinger non vede nessuna raffica di richieste: niente 429.
                dimensione = file_locale.stat().st_size
                print(f"    carico il video dal PC ({dimensione / 1048576:.1f} MB)...")
                with open(file_locale, "rb") as fh:
                    r = requests.post(
                        url_upload,
                        headers={
                            "Authorization": f"OAuth {self.token}",
                            "offset": "0",
                            "file_size": str(dimensione),
                            "Content-Type": "application/octet-stream",
                        },
                        data=fh,
                        timeout=900,
                    )
                if r.status_code != 200:
                    raise RuntimeError(f"FB upload reel (dal PC): {r.text[:300]}")
            else:
                # Riserva: faccio scaricare il file dal sito, con ritentativi.
                attese = [0, 20, 45, 90, 180]
                for i, pausa in enumerate(attese):
                    if pausa:
                        print(f"    il server e' sotto sforzo, riprovo tra {pausa}s "
                              f"(tentativo {i + 1} di {len(attese)})...")
                        time.sleep(pausa)
                    r = requests.post(
                        url_upload,
                        headers={"Authorization": f"OAuth {self.token}",
                                 "file_url": video_url},
                        timeout=600,
                    )
                    if r.status_code == 200:
                        break
                    testo = r.text[:300]
                    if "429" not in testo and "Too Many Requests" not in testo:
                        raise RuntimeError(f"FB upload reel: {testo}")
                else:
                    raise RuntimeError(
                        "FB upload reel: il server continua a rispondere 429.")

        finish = {
            "upload_phase": "finish",
            "video_id": video_id,
            "description": description,
        }
        if when:
            finish["video_state"] = "SCHEDULED"
            finish["scheduled_publish_time"] = int(when.timestamp())
        else:
            finish["video_state"] = "PUBLISHED"
        return self._post(f"{self.page_id}/video_reels", finish)

    def story_photo(self, image_url):
        up = self._post(f"{self.page_id}/photos", {"url": image_url, "published": "false"})
        return self._post(f"{self.page_id}/photo_stories", {"photo_id": up["id"]})


# ---------------------------------------------------------------------------
# TIKTOK (Content Posting API)
# ---------------------------------------------------------------------------


class TikTokClient:
    """
    Manda i video nelle BOZZE di TikTok (inbox). Tu apri l'app e confermi.
    E' l'unico modo senza passare l'audit di TikTok, che richiede settimane.

    Il file viene scaricato dal sito e rimandato a TikTok come byte:
    cosi' non serve verificare il dominio nel portale sviluppatori.
    """

    def __init__(self, cfg):
        self.cfg = cfg
        self.sez = cfg.get("tiktok", {})
        self.dati = self._leggi_dati()
        self.token = self.dati.get("access_token", "")
        self.dry = False

    # --- gestione del token -------------------------------------------------

    def _percorso(self) -> Path:
        return Path(self.cfg["_dir"]) / "token_tiktok.json"

    def _leggi_dati(self) -> dict:
        da_ambiente = os.environ.get("TOKEN_TIKTOK", "").strip()
        if da_ambiente:
            try:
                return json.loads(da_ambiente)
            except Exception:
                return {"access_token": da_ambiente}
        f = self._percorso()
        if f.exists():
            try:
                return json.loads(f.read_text(encoding="utf-8"))
            except Exception:
                return {}
        return {}

    def _salva_dati(self, dati: dict) -> None:
        self._percorso().write_text(
            json.dumps(dati, indent=2), encoding="utf-8")

    def rinnova(self) -> bool:
        """Il token TikTok dura 24 ore: si rinnova col refresh_token."""
        refresh = self.dati.get("refresh_token")
        client_key = self.sez.get("client_key")
        client_secret = self.sez.get("client_secret")
        if not (refresh and client_key and client_secret):
            return False

        r = requests.post("https://open.tiktokapis.com/v2/oauth/token/", data={
            "client_key": client_key,
            "client_secret": client_secret,
            "grant_type": "refresh_token",
            "refresh_token": refresh,
        }, headers={"Content-Type": "application/x-www-form-urlencoded"},
            timeout=60).json()

        if "access_token" not in r:
            print(f"    rinnovo TikTok non riuscito: {r.get('error_description', r)}")
            return False

        self.dati.update({
            "access_token": r["access_token"],
            "refresh_token": r.get("refresh_token", refresh),
            "rinnovato_il": datetime.now().isoformat(timespec="seconds"),
        })
        self.token = r["access_token"]
        self._salva_dati(self.dati)
        print("    token TikTok rinnovato")
        return True

    def _headers(self):
        return {
            "Authorization": f"Bearer {self.token}",
            "Content-Type": "application/json; charset=UTF-8",
        }

    def verifica(self) -> Optional[str]:
        """Controlla il token e, se scaduto, prova a rinnovarlo da solo."""
        if not self.token:
            return ("Nessun token TikTok. Lancia:\n"
                    "     py ddp_publisher.py token-tiktok --config config.json")
        for tentativo in (1, 2):
            r = requests.get("https://open.tiktokapis.com/v2/user/info/",
                             params={"fields": "open_id,display_name"},
                             headers={"Authorization": f"Bearer {self.token}"},
                             timeout=60).json()
            if "error" in r and r["error"].get("code") not in ("ok", None):
                if tentativo == 1 and self.rinnova():
                    continue
                return f"Token TikTok rifiutato: {r['error'].get('message')}"
            nome = r.get("data", {}).get("user", {}).get("display_name", "?")
            print(f"TikTok collegato: {nome}")
            return None
        return "Token TikTok non utilizzabile."

    # --- pubblicazione ------------------------------------------------------

    def publish_video(self, video_url, title, file_locale: Optional[Path] = None):
        if self.dry:
            print(f"    POST {TIKTOK}/post/publish/inbox/video/init/  "
                  f"(FILE_UPLOAD, titolo: {title[:60]}...)")
            return {"data": {"publish_id": "DRYRUN"}}

        # 1. prendo i byte del video: dal disco se c'e', altrimenti dal sito
        if file_locale and file_locale.exists():
            dati_video = file_locale.read_bytes()
        else:
            print("    scarico il video dal sito...")
            risposta = requests.get(video_url, timeout=300)
            if risposta.status_code != 200:
                raise RuntimeError(
                    f"TikTok: non riesco a scaricare il video "
                    f"(HTTP {risposta.status_code})")
            dati_video = risposta.content

        dimensione = len(dati_video)
        print(f"    invio a TikTok ({dimensione / 1048576:.1f} MB)...")

        # 2. apro il caricamento (un pezzo solo: i video sono piccoli)
        init = requests.post(
            f"{TIKTOK}/post/publish/inbox/video/init/",
            headers=self._headers(),
            data=json.dumps({"source_info": {
                "source": "FILE_UPLOAD",
                "video_size": dimensione,
                "chunk_size": dimensione,
                "total_chunk_count": 1,
            }}),
            timeout=120).json()

        errore = init.get("error", {})
        if errore.get("code") not in ("ok", None):
            raise RuntimeError(f"TikTok: {errore.get('message')}")

        upload_url = init["data"]["upload_url"]
        publish_id = init["data"]["publish_id"]

        # 3. mando i byte
        put = requests.put(
            upload_url,
            headers={
                "Content-Range": f"bytes 0-{dimensione - 1}/{dimensione}",
                "Content-Type": "video/mp4",
            },
            data=dati_video,
            timeout=900)
        if put.status_code not in (200, 201, 204):
            raise RuntimeError(f"TikTok caricamento: {put.text[:300]}")

        return {"data": {"publish_id": publish_id}}


# ---------------------------------------------------------------------------
# ESECUZIONE
# ---------------------------------------------------------------------------


def publish_task(cfg, t: Task, clients, dry: bool) -> str:
    urls = [media_url(cfg, m) for m in t.media]
    cover = media_url(cfg, t.cover) if t.cover else None
    caption = t.caption

    if t.platform == "instagram":
        ig = clients["ig"]
        if t.fmt == "reel":
            r = ig.reel(urls[0], caption, cover)
        elif t.fmt == "carosello":
            r = ig.carousel(urls, caption)
        elif t.fmt == "storia":
            r = ig.story(urls[0])
            # Lo sticker link va aggiunto a mano: l'API non lo espone.
        else:
            raise RuntimeError(f"formato IG non gestito: {t.fmt}")
        return f"ig media {r.get('id')}"

    if t.platform == "facebook":
        fb = clients["fb"]
        if t.fmt == "reel":
            locale = Path(cfg["packs_root"]) / t.media[0]
            r = fb.reel(urls[0], caption, file_locale=locale)
        elif t.fmt == "carosello":
            r = fb.carousel(urls, caption)
        elif t.fmt == "storia":
            r = fb.story_photo(urls[0])
        else:
            raise RuntimeError(f"formato FB non gestito: {t.fmt}")
        return f"fb {r.get('id') or r.get('post_id')}"

    if t.platform == "tiktok":
        tk = clients["tk"]
        locale = Path(cfg["packs_root"]) / t.media[0]
        r = tk.publish_video(urls[0], caption, file_locale=locale)
        pid = r.get("data", {}).get("publish_id")
        return f"tiktok {pid} (in bozza - conferma dall'app)"

    raise RuntimeError(f"piattaforma sconosciuta: {t.platform}")


PLAT_KEY = {"instagram": "ig", "facebook": "fb", "tiktok": "tk"}


def filter_tasks(tasks: List[Task], args) -> List[Task]:
    """Applica i filtri da riga di comando: piattaforma, libro, contenuto, task."""
    out = tasks

    if getattr(args, "only", None):
        wanted = {p.strip().lower() for p in args.only.split(",") if p.strip()}
        unknown = wanted - set(PLAT_KEY)
        if unknown:
            sys.exit(f"Piattaforma sconosciuta: {', '.join(unknown)}. "
                     f"Valori validi: instagram, facebook, tiktok")
        out = [t for t in out if t.platform in wanted]

    if getattr(args, "book", None):
        b = args.book.lower()
        out = [t for t in out if b in t.book.lower()]

    if getattr(args, "content", None):
        wanted = {c.strip().upper() for c in args.content.split(",") if c.strip()}
        out = [t for t in out if t.content_id.upper() in wanted]

    if getattr(args, "fmt", None):
        wanted = {f.strip().lower() for f in args.fmt.split(",") if f.strip()}
        out = [t for t in out if t.fmt in wanted]

    if getattr(args, "task", None):
        wanted = {x.strip().upper() for x in args.task.split(",") if x.strip()}
        out = [t for t in out if t.task_id.upper() in wanted]

    return out


def make_clients(cfg, dry: bool):
    clients = {}
    if cfg.get("instagram", {}).get("enabled"):
        c = InstagramClient(cfg); c.dry = dry; clients["ig"] = c
    if cfg.get("facebook", {}).get("enabled"):
        c = FacebookClient(cfg); c.dry = dry; clients["fb"] = c
    if cfg.get("tiktok", {}).get("enabled"):
        c = TikTokClient(cfg); c.dry = dry; clients["tk"] = c
    return clients


def cmd_run(cfg, args) -> None:
    dry = args.dry_run
    tasks = filter_tasks(load_queue(cfg), args)
    state = load_state(cfg)
    clients = make_clients(cfg, dry)

    # Controllo i token PRIMA. Se uno e' scaduto salto solo quella piattaforma:
    # un problema su Facebook non deve bloccare Instagram, e viceversa.
    if not dry:
        if "fb" in clients:
            guaio = verifica_token_fb(cfg)
            if guaio:
                print(f"FACEBOOK saltato - token non valido: {guaio.splitlines()[0]}")
                clients.pop("fb")
        if "ig" in clients:
            guaio = verifica_token_ig(cfg)
            if guaio:
                print(f"INSTAGRAM saltato - token non valido: {guaio.splitlines()[0]}")
                clients.pop("ig")
            elif not getattr(args, "no_sync", False):
                # riallineo col profilo: se un contenuto e' gia' online, non lo rifaccio
                cmd_ig_sync(cfg)
                state = load_state(cfg)
        if "tk" in clients:
            guaio = clients["tk"].verifica()
            if guaio:
                print(f"TIKTOK saltato - {guaio.splitlines()[0]}")
                clients.pop("tk")
        if not clients:
            print("Nessuna piattaforma utilizzabile. Rigenera i token.")
            return
    now = datetime.now()
    lower = now - timedelta(minutes=args.window)

    due = [
        t for t in tasks
        if lower <= t.dt <= now
        and t.key not in state["done"]
        and PLAT_KEY.get(t.platform) in clients
    ]

    if not due:
        print(f"[{now:%d/%m %H:%M}] Niente da pubblicare.")
        skipped = [t for t in tasks if lower <= t.dt <= now
                   and PLAT_KEY.get(t.platform) not in clients]
        if skipped:
            plats = sorted({t.platform for t in skipped})
            print(f"  ({len(skipped)} task scaduti su {', '.join(plats)}, "
                  f"ma quelle piattaforme sono disattivate o escluse dai filtri)")
        return

    for t in due:
        print(f"[{now:%H:%M}] {t.platform}/{t.fmt} {t.content_id} - {t.title}")
        try:
            info = publish_task(cfg, t, clients, dry)
            print(f"    OK: {info}")
            if not dry:
                mark(state, t, True, info)
        except Exception as e:
            print(f"    ERRORE: {e}")
            if not dry:
                mark(state, t, False, str(e))
        if not dry:
            save_state(cfg, state)
            time.sleep(3)


def verifica_token_fb(cfg) -> Optional[str]:
    """
    Prima di pubblicare, controlla che il token di Facebook sia davvero
    quello della Pagina. Restituisce un messaggio d'errore, o None se e' a posto.
    """
    fb = cfg.get("facebook", {})
    token = leggi_token(cfg, "facebook")
    page_id = str(fb.get("page_id") or "").strip()

    if not token or "INCOLLA" in token.upper():
        return "Il token di Facebook non e' stato inserito nel config."
    if any(c in token for c in "\r\n \t"):
        return ("Il token contiene spazi o a capo: si e' rotto durante la copia.\n"
                "     Riprendilo con 'py controlla_token.py' e incollalo tutto di seguito.")

    try:
        r = requests.get(f"{GRAPH}/me",
                         params={"fields": "id,name", "access_token": token},
                         timeout=30).json()
    except Exception as e:
        return f"Non riesco a contattare Facebook: {e}"

    if "error" in r:
        return f"Token rifiutato: {r['error'].get('message')}"

    if str(r.get("id")) != page_id:
        return (f"Il token e' di '{r.get('name')}' (id {r.get('id')}),\n"
                f"     ma nel config la Pagina e' {page_id}.\n"
                f"     Serve il TOKEN DELLA PAGINA. Lancia: py controlla_token.py")

    return None


def cmd_test(cfg, args) -> None:
    """
    Pubblica UN SOLO contenuto, adesso, ignorando la data del calendario.
    Serve per la prima prova: vedi se arriva davvero prima di lanciare tutto.

    Esempio:
      py ddp_publisher.py test --config config.json --only facebook --content TRAILER
    """
    tasks = filter_tasks(load_queue(cfg), args)
    if not tasks:
        sys.exit("Nessun task corrisponde ai filtri. Prova 'plan' per vedere cosa c'e'.")

    clients = make_clients(cfg, args.dry_run)
    tasks = [t for t in tasks if PLAT_KEY.get(t.platform) in clients]
    if not tasks:
        sys.exit("I task trovati sono su piattaforme disattivate nel config.")

    t = tasks[0]

    if t.platform == "instagram" and not args.dry_run:
        guaio = verifica_token_ig(cfg)
        if guaio:
            print(f"\n>>> CONTROLLO TOKEN INSTAGRAM FALLITO:\n     {guaio}")
            print("\nNon pubblico niente.")
            return

    if t.platform == "facebook" and not args.dry_run:
        problema = verifica_token_fb(cfg)
        if problema:
            print(f"\n>>> CONTROLLO TOKEN FALLITO:\n     {problema}")
            print("\nNon pubblico niente finche' il token non e' corretto.")
            return
        print("Token della Pagina: verificato.")

    print("\n--- PROVA SINGOLA ---")
    print(f"Piattaforma : {t.platform}")
    print(f"Formato     : {t.fmt}")
    print(f"Libro       : {t.book}  ({t.content_id} - {t.title})")
    print(f"File        :")
    for m in t.media:
        print(f"              {media_url(cfg, m)}")
    if t.cover:
        print(f"Copertina   : {media_url(cfg, t.cover)}")
    print(f"Caption     : {t.caption[:180].replace(chr(10), ' / ')}...")
    print(f"\nATTENZIONE: questo contenuto viene pubblicato ORA, davvero.")
    if len(tasks) > 1:
        print(f"(altri {len(tasks) - 1} task corrispondono ai filtri, ma ne pubblico solo uno)")

    if not args.dry_run and not args.yes:
        risposta = input("\nScrivi SI e premi Invio per procedere: ").strip().upper()
        if risposta not in ("SI", "SÌ", "S", "YES"):
            print("Annullato. Non e' stato pubblicato niente.")
            return

    state = load_state(cfg)
    try:
        info = publish_task(cfg, t, clients, args.dry_run)
        print(f"\nOK: {info}")
        print("Vai a controllare sul profilo. Se lo vedi, funziona tutto.")
        if not args.dry_run:
            mark(state, t, True, info)
            save_state(cfg, state)
    except Exception as e:
        print(f"\nERRORE: {e}")
        print("\nSuggerimenti:")
        print("  - 'media_base_url' nel config e' l'indirizzo giusto?")
        print("  - i file si aprono dal browser? (provali a mano)")
        print("  - il token e' ancora valido?")


def cmd_export(cfg, args) -> None:
    """
    Non pubblica niente: prepara una cartella con tutto pronto da copiare a mano.
    Utile finche' gli account non sono collegati, o per le Storie.
    """
    tasks = filter_tasks(load_queue(cfg), args)
    now = datetime.now()
    limit = now + timedelta(days=args.days)
    tasks = [t for t in tasks if now - timedelta(days=2) <= t.dt <= limit]

    if not tasks:
        print("Nessun contenuto nel periodo indicato.")
        return

    outdir = Path(cfg["_dir"]) / "DA_PUBBLICARE_A_MANO"
    outdir.mkdir(exist_ok=True)
    packs_root = Path(cfg["packs_root"])

    righe = ["# Da pubblicare a mano", ""]
    for t in sorted(tasks, key=lambda x: x.dt):
        righe.append(f"## {t.dt:%d/%m/%Y ore %H:%M}  -  {t.platform.upper()} / {t.fmt}")
        righe.append(f"**{t.book} - {t.content_id}: {t.title}**")
        righe.append("")
        righe.append("File da caricare (dal PC):")
        for m in t.media:
            righe.append(f"- `{packs_root / m}`")
        if t.cover:
            righe.append(f"- copertina: `{packs_root / t.cover}`")
        righe.append("")
        righe.append("Oppure scaricali dal telefono (tieni premuto -> Salva):")
        for m in t.media:
            righe.append(f"- {media_url(cfg, m)}")
        if t.cover:
            righe.append(f"- copertina: {media_url(cfg, t.cover)}")
        if t.fmt == "storia" and t.link:
            righe.append("")
            righe.append(f"Sticker link: {t.link}")
        if t.caption:
            righe.append("")
            righe.append("Caption da copiare:")
            righe.append("")
            righe.append("```")
            righe.append(t.caption)
            righe.append("```")
        righe.append("")
        righe.append("---")
        righe.append("")

    f = outdir / f"da_pubblicare_{now:%Y-%m-%d}.md"
    f.write_text("\n".join(righe), encoding="utf-8")
    print(f"Creato: {f}")
    print(f"{len(tasks)} contenuti, prossimi {args.days} giorni.")
    print("Aprilo, copia le caption da li'. Niente e' stato pubblicato.")


def cmd_token(cfg, args) -> None:
    """
    Prende il token della Pagina e lo salva in token_facebook.txt.
    Tu incolli solo il token utente dell'Explorer: al resto pensa il programma,
    cosi' non devi mai copiare il token lungo dalla finestra nera.
    """
    page_id = str(cfg.get("facebook", {}).get("page_id") or "").strip()
    if not page_id:
        sys.exit("Metti prima 'page_id' nel config.json.")

    print("\n--- RECUPERO TOKEN DELLA PAGINA ---\n")
    print("1. Vai su developers.facebook.com/tools/explorer")
    print("2. In 'Utente o Pagina' lascia pure 'Token utente'")
    print("3. Clicca 'Generate Access Token' e autorizza")
    print("4. Copia il token dal campo in alto (icona dei due fogli)")
    print("\nPoi incollalo qui sotto con il TASTO DESTRO del mouse e premi Invio.")
    print("(Il testo non si vede tutto: e' normale, va bene lo stesso.)\n")

    user_token = input("Token utente: ").strip().replace(" ", "").replace("\n", "")
    if not user_token:
        print("Niente token, annullo.")
        return

    # Se ci sono ID app e chiave segreta, allungo prima il token utente:
    # cosi' il token della Pagina che ne deriva non scade mai.
    app_id = str(cfg.get("facebook", {}).get("app_id") or "").strip()
    app_secret = str(cfg.get("facebook", {}).get("app_secret") or "").strip()
    if app_id and app_secret:
        print("\nAllungo il token utente...")
        r = requests.get(f"{GRAPH}/oauth/access_token", params={
            "grant_type": "fb_exchange_token",
            "client_id": app_id,
            "client_secret": app_secret,
            "fb_exchange_token": user_token,
        }, timeout=30).json()
        if "access_token" in r:
            user_token = r["access_token"]
            print("  fatto: il token della Pagina sara' permanente.")
        else:
            msg = r.get("error", {}).get("message", "?")
            print(f"  non riuscito ({msg}) - vado avanti col token breve.")

    print("\nCerco le tue Pagine...")
    r = requests.get(f"{GRAPH}/me/accounts", params={
        "fields": "name,id,access_token",
        "access_token": user_token,
    }, timeout=30).json()

    if "error" in r:
        print(f"\nERRORE: {r['error'].get('message')}")
        print("Il token utente e' scaduto o incollato male. Riprova.")
        return

    pagine = r.get("data", [])
    if not pagine:
        print("\nNessuna Pagina trovata con questo token.")
        print("Nell'Explorer servono i permessi pages_show_list e pages_manage_posts.")
        return

    scelta = next((p for p in pagine if str(p.get("id")) == page_id), None)
    if scelta is None:
        print(f"\nLa Pagina {page_id} non e' tra quelle autorizzate. Trovate:")
        for p in pagine:
            print(f"  - {p.get('name')}  (id {p.get('id')})")
        return

    f = Path(cfg["_dir"]) / "token_facebook.txt"
    f.write_text(scelta["access_token"], encoding="utf-8")

    print(f"\n*** FATTO ***")
    print(f"Token della Pagina '{scelta.get('name')}' salvato in:")
    print(f"  {f}")
    print("\nNon devi copiare niente: il programma lo legge da li'.")
    print("Controlla con:  py ddp_publisher.py token-check --config config.json")


def verifica_token_ig(cfg) -> Optional[str]:
    """Controlla il token di Instagram (accesso diretto)."""
    sez = cfg.get("instagram", {})
    token = leggi_token(cfg, "instagram")
    if not token:
        return ("Nessun token Instagram trovato.\n"
                "     Serve il file token_instagram.txt in C:\\DDP")
    if any(c in token for c in "\r\n \t"):
        return "Il token Instagram contiene spazi o a capo: ricopialo su una riga sola."

    base = ("https://graph.instagram.com" if sez.get("instagram_login") else GRAPH)
    try:
        r = requests.get(f"{base}/me",
                         params={"fields": "id,username,account_type",
                                 "access_token": token},
                         timeout=30).json()
    except Exception as e:
        return f"Non riesco a contattare Instagram: {e}"

    if "error" in r:
        return f"Token rifiutato: {r['error'].get('message')}"

    print(f"Instagram collegato: @{r.get('username')} "
          f"({r.get('account_type')}, id {r.get('id')})")
    return None


def cmd_token_check(cfg) -> int:
    """Restituisce 0 se il token va bene, 1 altrimenti (serve ai file .cmd)."""
    if cfg.get("instagram", {}).get("enabled"):
        guaio = verifica_token_ig(cfg)
        if guaio:
            print(f"\n>>> INSTAGRAM: {guaio}")
        else:
            print("*** Token Instagram: OK ***")

    if not cfg.get("facebook", {}).get("enabled"):
        return 0

    problema = verifica_token_fb(cfg)
    if problema:
        print(f"\n>>> {problema}")
        print("\nPer sistemare:  py ddp_publisher.py token --config config.json")
        return 1
    print("\n*** TOKEN CORRETTO: e' quello della Pagina. Puoi pubblicare. ***")
    return 0


def cmd_token_instagram(cfg, args) -> None:
    """
    Collega Instagram SENZA passare dalla Pagina Facebook.
    Usa l'accesso diretto di Instagram: serve solo un account professionale.
    """
    sez = cfg.get("instagram", {})
    app_id = str(sez.get("instagram_app_id") or "").strip()
    app_secret = str(sez.get("instagram_app_secret") or "").strip()
    redirect = str(sez.get("redirect_uri") or "").strip()

    segnaposto = ("IL_NUMERO", "LA_CHIAVE", "IL_TUO", "LA_TUA", "...", "INCOLLA")
    for valore, etichetta in ((app_id, "instagram_app_id"),
                              (app_secret, "instagram_app_secret")):
        if valore and any(s in valore.upper() for s in
                          (x.upper() for x in segnaposto)):
            print(f"\nNel config, \"{etichetta}\" contiene ancora il testo di")
            print(f"esempio: {valore}")
            print("\nSostituiscilo con il valore vero preso dall'app Instagram.")
            print("  ID app Instagram      -> instagram_app_id")
            print("  Chiave segreta        -> instagram_app_secret")
            return

    if not (app_id and app_secret and redirect):
        print("\nManca qualcosa nel config.json, sezione \"instagram\":")
        print('  "instagram_login": true,')
        print('  "instagram_app_id": "...",')
        print('  "instagram_app_secret": "...",')
        print('  "redirect_uri": "https://diaridipelle.it/"')
        print("\nI due valori si prendono da developers.facebook.com/apps ->")
        print("la tua app -> Instagram -> Configurazione API con accesso Instagram.")
        print("Sono DIVERSI dall'ID e dalla chiave segreta dell'app Facebook.")
        return

    scopes = "instagram_business_basic,instagram_business_content_publish"
    url = "https://www.instagram.com/oauth/authorize?" + urlencode({
        "client_id": app_id,
        "redirect_uri": redirect,
        "scope": scopes,
        "response_type": "code",
    })

    print("\n--- COLLEGAMENTO INSTAGRAM (senza Pagina Facebook) ---\n")
    print("1. Copia questo indirizzo e aprilo nel browser:\n")
    print(f"   {url}\n")
    print("2. Accedi con l'account @diarididipelle e autorizza")
    print("3. Il browser ti rimanda al tuo sito. Sembrera' una pagina normale,")
    print("   ma nella BARRA DEGLI INDIRIZZI in alto ci sara' scritto")
    print("   qualcosa come:  https://diaridipelle.it/?code=AQBx...")
    print("4. Copia TUTTO l'indirizzo dalla barra e incollalo qui sotto.\n")

    risposta = input("Indirizzo (o solo il codice): ").strip()
    if not risposta:
        print("Annullato.")
        return

    if "oauth/authorize" in risposta or "client_id=" in risposta:
        print("\nQuesto e' l'indirizzo che ti ho DATO io, non quello di ritorno.")
        print("Devi prima aprirlo nel browser e autorizzare. Poi il browser")
        print("ti porta su diaridipelle.it e SOLO ALLORA nella barra degli")
        print("indirizzi compare qualcosa con ?code=  dentro. Quello mi serve.")
        return

    codice = risposta
    if "code=" in risposta:
        codice = risposta.split("code=", 1)[1].split("&")[0]
    codice = codice.rstrip("#_").strip()
    print(f"\n  codice letto      : {codice[:25]}... ({len(codice)} caratteri)")
    print(f"  redirect_uri usato: {redirect}")
    print("  Questo indirizzo deve essere IDENTICO a quello registrato in")
    print("  Instagram -> Configura Instagram Business Login -> Configura.")

    print("\nScambio il codice con un token...")
    risp = requests.post("https://api.instagram.com/oauth/access_token", data={
        "client_id": app_id,
        "client_secret": app_secret,
        "grant_type": "authorization_code",
        "redirect_uri": redirect,
        "code": codice,
    }, timeout=60)
    try:
        r = risp.json()
    except Exception:
        print(f"\nInstagram ha risposto in modo inatteso (HTTP {risp.status_code}):")
        print(f"  {risp.text[:400]}")
        print("\nControlla ID app e chiave segreta nel config.")
        return

    if "access_token" not in r:
        print(f"\nERRORE: {r}")
        print("\nCause tipiche: il codice si usa una volta sola (rigenera l'indirizzo),")
        print("oppure il redirect_uri non combacia con quello scritto nell'app.")
        return

    breve = r["access_token"]
    ig_id = r.get("user_id")
    print(f"  token ottenuto (utente {ig_id})")

    print("Lo allungo a 60 giorni...")
    r2 = requests.get("https://graph.instagram.com/access_token", params={
        "grant_type": "ig_exchange_token",
        "client_secret": app_secret,
        "access_token": breve,
    }, timeout=60).json()

    token = r2.get("access_token", breve)
    if "access_token" in r2:
        giorni = int(r2.get("expires_in", 0)) // 86400
        print(f"  fatto: dura {giorni} giorni")
    else:
        print(f"  non riuscito: uso il token breve (dura un'ora)")

    f = Path(cfg["_dir"]) / "token_instagram.txt"
    f.write_text(token, encoding="utf-8")

    me = requests.get("https://graph.instagram.com/me", params={
        "fields": "id,username,account_type",
        "access_token": token,
    }, timeout=60).json()

    print(f"\n*** FATTO ***")
    print(f"Account collegato : @{me.get('username')} ({me.get('account_type')})")
    print(f"Token salvato in  : {f}")
    print("\nNel config.json, sezione instagram, metti:")
    print('  "enabled": true,')
    print('  "instagram_login": true')
    print("\nPoi prova:")
    print("  py ddp_publisher.py test --config config.json --only instagram "
          "--content TRAILER --fmt carosello --dry-run")


def cmd_reset(cfg, args) -> None:
    """
    Toglie dal registro dei contenuti gia' pubblicati, cosi' possono
    uscire di nuovo. Serve per riproporre vecchi contenuti.

    Esempi:
      py ddp_publisher.py reset --config config.json --content TRAILER
      py ddp_publisher.py reset --config config.json --book Notti --only instagram
    """
    if not any(getattr(args, f, None) for f in
               ("content", "book", "only", "fmt", "task")):
        print("Serve almeno un filtro, per non azzerare tutto per sbaglio.")
        print("  --content TRAILER     --book Notti")
        print("  --only instagram      --fmt reel        --task T009")
        return

    tasks = filter_tasks(load_queue(cfg), args)
    state = load_state(cfg)

    da_togliere = [t for t in tasks
                   if t.key in state["done"] or t.key in state["failed"]]

    if not da_togliere:
        print("Nessun contenuto gia' pubblicato corrisponde ai filtri.")
        return

    print(f"\n{len(da_togliere)} contenuti tornerebbero 'da pubblicare':\n")
    for t in da_togliere:
        stato = "pubblicato" if t.key in state["done"] else "fallito"
        print(f"  {t.dt:%d/%m %H:%M}  {t.platform:9} {t.fmt:9} "
              f"{t.content_id:8} ({stato})")

    print("\nUsciranno di nuovo appena arriva la loro data.")
    if not args.yes:
        risposta = input("\nScrivi SI per confermare: ").strip().upper()
        if risposta not in ("SI", "SÌ", "S", "YES"):
            print("Annullato.")
            return

    for t in da_togliere:
        state["done"].pop(t.key, None)
        state["failed"].pop(t.key, None)
    save_state(cfg, state)
    print(f"\nFatto: {len(da_togliere)} contenuti rimessi in coda.")
    print("Ricordati di ricaricare state.json su GitHub.")


def cmd_ig_refresh(cfg) -> None:
    """
    Allunga il token di Instagram di altri 60 giorni.
    Va fatto prima della scadenza: se lanciato ogni settimana, il token
    non scade mai piu'.
    """
    token = leggi_token(cfg, "instagram")
    if not token:
        print("Nessun token Instagram da rinnovare.")
        return

    r = requests.get("https://graph.instagram.com/refresh_access_token", params={
        "grant_type": "ig_refresh_token",
        "access_token": token,
    }, timeout=60).json()

    if "access_token" not in r:
        print(f"Rinnovo non riuscito: {r}")
        print("Se il token e' gia' scaduto, rifai:")
        print("  py ddp_publisher.py token-instagram --config config.json")
        return

    nuovo = r["access_token"]
    giorni = int(r.get("expires_in", 0)) // 86400
    f = Path(cfg["_dir"]) / "token_instagram.txt"
    f.write_text(nuovo, encoding="utf-8")
    print(f"Token Instagram rinnovato: valido altri {giorni} giorni.")
    print(f"Salvato in {f}")

    # Su GitHub il file non si puo' riusare: stampo il valore per il segreto.
    if os.environ.get("GITHUB_ACTIONS"):
        print("::add-mask::" + nuovo)
        print("ATTENZIONE: aggiorna il segreto TOKEN_INSTAGRAM su GitHub.")


def cmd_ig_sync(cfg) -> None:
    """
    Chiede a Instagram cosa e' gia' stato pubblicato sul profilo e aggiorna
    il registro locale. Cosi' non ripubblichi cose gia' uscite, anche se
    state.json si perde o se hai pubblicato a mano dall'app.
    """
    guaio = verifica_token_ig(cfg)
    if guaio:
        print(f"\n>>> {guaio}")
        return

    sez = cfg.get("instagram", {})
    token = leggi_token(cfg, "instagram")
    base = "https://graph.instagram.com" if sez.get("instagram_login") else GRAPH
    chi = "me" if sez.get("instagram_login") else sez.get("ig_user_id", "me")

    pubblicati = []
    url = f"{base}/{chi}/media"
    params = {"fields": "id,caption,media_type,timestamp", "limit": 100,
              "access_token": token}
    for _ in range(10):                     # scorro le pagine
        r = requests.get(url, params=params, timeout=60).json()
        if "error" in r:
            print(f"Instagram: {r['error'].get('message')}")
            break
        pubblicati.extend(r.get("data", []))
        prossima = r.get("paging", {}).get("next")
        if not prossima:
            break
        url, params = prossima, {}

    if not pubblicati:
        print("Nessun contenuto trovato sul profilo. Registro invariato.")
        return

    print(f"Sul profilo risultano {len(pubblicati)} contenuti pubblicati.")

    def impronta(testo: str) -> str:
        """Prime parole della caption, senza spazi ne' maiuscole: serve
        a riconoscere lo stesso contenuto anche se Instagram lo riformatta."""
        pulito = re.sub(r"\s+", " ", (testo or "")).strip().lower()
        return pulito[:80]

    # Il tipo conta: Reel e carosello dello stesso capitolo hanno la STESSA
    # caption, quindi confrontare solo il testo li confonderebbe.
    def tipo_ig(media_type: str) -> str:
        mt = (media_type or "").upper()
        if mt in ("CAROUSEL_ALBUM", "CAROUSEL"):
            return "carosello"
        if mt in ("VIDEO", "REELS", "REEL"):
            return "reel"
        return "immagine"

    gia_online = {(impronta(m.get("caption")), tipo_ig(m.get("media_type")))
                  for m in pubblicati if m.get("caption")}

    tasks = load_queue(cfg)
    state = load_state(cfg)
    aggiunti = 0

    for t in tasks:
        if t.platform != "instagram" or t.key in state["done"]:
            continue
        if not t.caption:
            continue
        atteso = "carosello" if t.fmt == "carosello" else "reel"
        if (impronta(t.caption), atteso) in gia_online:
            mark(state, t, True, "ig gia' pubblicato (trovato sul profilo)")
            aggiunti += 1
            print(f"  gia' online: {t.fmt:9} {t.content_id} - {t.title}")

    save_state(cfg, state)
    print(f"\nRegistro aggiornato: {aggiunti} contenuti segnati come gia' fatti.")
    if aggiunti == 0:
        print("(nessuna corrispondenza: nessun rischio di doppioni)")


def cmd_token_tiktok(cfg, args) -> None:
    """Collega l'account TikTok. Da fare una volta sola."""
    sez = cfg.get("tiktok", {})
    client_key = str(sez.get("client_key") or "").strip()
    client_secret = str(sez.get("client_secret") or "").strip()
    redirect = str(sez.get("redirect_uri") or "").strip()

    if not (client_key and client_secret and redirect):
        print("\nManca qualcosa nel config.json, sezione \"tiktok\":")
        print('  "enabled": true,')
        print('  "client_key": "...",')
        print('  "client_secret": "...",')
        print('  "redirect_uri": "https://diaridipelle.it/"')
        print("\nI valori stanno su developers.tiktok.com -> la tua app ->")
        print("Basic information (Client key e Client secret).")
        return

    stato = "ddp" + str(int(time.time()))
    url = "https://www.tiktok.com/v2/auth/authorize/?" + urlencode({
        "client_key": client_key,
        "scope": "user.info.basic,video.upload",
        "response_type": "code",
        "redirect_uri": redirect,
        "state": stato,
    })

    print("\n--- COLLEGAMENTO TIKTOK ---\n")
    print("1. Copia questo indirizzo e aprilo nel browser:\n")
    print(f"   {url}\n")
    print("2. Accedi con l'account TikTok di Diari di Pelle e autorizza")
    print("3. Il browser ti rimanda a diaridipelle.it: copia dalla barra")
    print("   degli indirizzi tutto cio' che contiene ?code=\n")

    risposta = input("Indirizzo (o solo il codice): ").strip()
    if not risposta:
        print("Annullato.")
        return
    if "auth/authorize" in risposta:
        print("\nQuesto e' l'indirizzo che ti ho dato io, non quello di ritorno.")
        return

    codice = risposta
    if "code=" in risposta:
        codice = risposta.split("code=", 1)[1].split("&")[0]
    codice = unquote(codice.rstrip("#_").strip())

    print("\nScambio il codice con un token...")
    r = requests.post("https://open.tiktokapis.com/v2/oauth/token/", data={
        "client_key": client_key,
        "client_secret": client_secret,
        "code": codice,
        "grant_type": "authorization_code",
        "redirect_uri": redirect,
    }, headers={"Content-Type": "application/x-www-form-urlencoded"},
        timeout=60).json()

    if "access_token" not in r:
        print(f"\nERRORE: {r}")
        print("\nIl codice si usa una volta sola: rilancia il comando.")
        print("Controlla anche che redirect_uri combaci con quello nell'app.")
        return

    dati = {
        "access_token": r["access_token"],
        "refresh_token": r.get("refresh_token", ""),
        "open_id": r.get("open_id", ""),
        "creato_il": datetime.now().isoformat(timespec="seconds"),
    }
    f = Path(cfg["_dir"]) / "token_tiktok.json"
    f.write_text(json.dumps(dati, indent=2), encoding="utf-8")

    print("\n*** FATTO ***")
    print(f"Token salvato in: {f}")
    print("\nL'accesso dura 24 ore ma si rinnova da solo (il rinnovo vale un anno).")
    print("\nProva con:")
    print("  py ddp_publisher.py test --config config.json --only tiktok "
          "--content TRAILER --fmt reel --dry-run")


def cmd_tiktok_batch(cfg, args) -> None:
    """
    Manda in blocco i video nelle bozze di TikTok e prepara un file con
    le caption nello stesso ordine, pronte da copiare.

    L'API dell'inbox non accetta il testo: la caption va incollata a mano
    nell'app. Questo comando ti evita almeno di cercarla ogni volta.
    """
    if not cfg.get("tiktok", {}).get("enabled"):
        sys.exit("TikTok non abilitato nel config.")

    tk = TikTokClient(cfg)
    tk.dry = args.dry_run
    if not args.dry_run:
        guaio = tk.verifica()
        if guaio:
            print(f"\n>>> {guaio}")
            return

    tasks = filter_tasks(load_queue(cfg), args)
    state = load_state(cfg)
    da_fare = [t for t in tasks
               if t.platform == "tiktok" and t.key not in state["done"]]

    quanti = getattr(args, "limit", 0) or len(da_fare)
    da_fare = da_fare[:quanti]

    if not da_fare:
        print("Niente da caricare su TikTok.")
        return

    print(f"\n{len(da_fare)} video da mandare nelle bozze di TikTok.")
    print("Le caption non entrano nelle bozze: te le preparo in un file.\n")

    righe = ["# Caption TikTok — nell'ordine in cui hai caricato i video", ""]
    righe.append("Le bozze in TikTok appaiono con la piu' recente in alto,")
    righe.append("quindi leggi questo elenco dal BASSO verso l'ALTO.")
    righe.append("")
    righe.append("---")
    righe.append("")

    ok = errori = 0
    for n, t in enumerate(da_fare, 1):
        url = media_url(cfg, t.media[0])
        locale = Path(cfg["packs_root"]) / t.media[0]
        print(f"[{n}/{len(da_fare)}] {t.dt:%d/%m} {t.content_id} - {t.title}")
        try:
            r = tk.publish_video(url, t.caption, file_locale=locale)
            pid = r.get("data", {}).get("publish_id")
            print(f"    OK: {pid}")
            ok += 1
            if not args.dry_run:
                mark(state, t, True, f"tiktok bozza {pid}")
                save_state(cfg, state)

            righe.append(f"## {n}. {t.content_id} — {t.title}")
            righe.append(f"*(data prevista: {t.dt:%d/%m/%Y})*")
            righe.append("")
            righe.append("```")
            righe.append(t.caption)
            righe.append("```")
            righe.append("")
        except Exception as e:
            print(f"    ERRORE: {e}")
            errori += 1
            if not args.dry_run:
                mark(state, t, False, str(e))
                save_state(cfg, state)
        if not args.dry_run and n < len(da_fare):
            time.sleep(5)

    outdir = Path(cfg["_dir"]) / "DA_PUBBLICARE_A_MANO"
    outdir.mkdir(exist_ok=True)
    f = outdir / f"caption_tiktok_{datetime.now():%Y-%m-%d_%H%M}.md"
    f.write_text("\n".join(righe), encoding="utf-8")

    print(f"\nCaricati: {ok}   Falliti: {errori}")
    print(f"Caption pronte in: {f}")
    print("\nApri TikTok, vai nelle bozze e per ognuna incolla la caption")
    print("corrispondente. L'elenco va letto dal basso verso l'alto.")


def cmd_fb_list(cfg) -> None:
    """Chiede a Facebook l'elenco di cio' che e' davvero in coda sulla Pagina."""
    problema = verifica_token_fb(cfg)
    if problema:
        print(f"\n>>> {problema}")
        return

    page_id = str(cfg["facebook"]["page_id"]).strip()
    token = leggi_token(cfg, "facebook")

    trovati = []
    for endpoint, etichetta in (("scheduled_posts", "post"), ("video_reels", "reel")):
        r = requests.get(f"{GRAPH}/{page_id}/{endpoint}", params={
            "fields": "id,scheduled_publish_time,message,description,title",
            "limit": 100,
            "access_token": token,
        }, timeout=60).json()
        if "error" in r:
            continue
        for v in r.get("data", []):
            quando = v.get("scheduled_publish_time")
            if not quando:
                continue
            testo = (v.get("message") or v.get("description")
                     or v.get("title") or "")
            trovati.append((int(quando), etichetta, testo.split("\n")[0][:60]))

    if not trovati:
        print("Non risulta niente in coda.")
        print("Se hai appena programmato, aspetta qualche minuto e riprova.")
        print("Puoi controllare anche da business.facebook.com -> Pianificazione.")
        return

    trovati.sort()
    print(f"\n{len(trovati)} contenuti programmati sulla Pagina:\n")
    for quando, tipo, testo in trovati:
        print(f"  {datetime.fromtimestamp(quando):%d/%m/%Y %H:%M}  {tipo:6}  {testo}")


def cmd_sync(cfg) -> None:
    """
    Chiede a Facebook cosa risulta gia' programmato e aggiorna il registro locale.
    Serve se state.json si e' perso o disallineato: cosi' non ricarichi doppioni.
    """
    problema = verifica_token_fb(cfg)
    if problema:
        print(f"\n>>> {problema}")
        return

    page_id = str(cfg["facebook"]["page_id"]).strip()
    token = leggi_token(cfg, "facebook")

    programmati = []
    for endpoint in ("scheduled_posts", "video_reels"):
        dopo = None
        for _ in range(10):          # scorro le pagine di risultati
            params = {"fields": "id,scheduled_publish_time", "limit": 100,
                      "access_token": token}
            if dopo:
                params["after"] = dopo
            r = requests.get(f"{GRAPH}/{page_id}/{endpoint}", params=params,
                             timeout=60).json()
            if "error" in r:
                break
            for v in r.get("data", []):
                if v.get("scheduled_publish_time"):
                    programmati.append(
                        (int(v["scheduled_publish_time"]), v.get("id")))
            dopo = r.get("paging", {}).get("cursors", {}).get("after")
            if not dopo or not r.get("data"):
                break

    if not programmati:
        print("Facebook non riporta niente in coda. Registro invariato.")
        return

    print(f"Facebook ha {len(programmati)} contenuti in coda.")

    tasks = load_queue(cfg)
    state = load_state(cfg)
    aggiunti = 0

    for orario, post_id in programmati:
        quando = datetime.fromtimestamp(orario)
        # Cerco il task con la stessa data e ora (tolleranza di 2 minuti)
        for t in tasks:
            if t.platform != "facebook" or t.key in state["done"]:
                continue
            if abs((t.dt - quando).total_seconds()) <= 120:
                mark(state, t, True, f"fb gia' programmato {post_id}")
                aggiunti += 1
                print(f"  segnato come fatto: {t.dt:%d/%m %H:%M} "
                      f"{t.fmt} {t.content_id}")
                break

    save_state(cfg, state)
    print(f"\nRegistro aggiornato: {aggiunti} contenuti segnati come gia' fatti.")
    if aggiunti == 0:
        print("(erano gia' tutti segnati: nessun rischio di doppioni)")


def cmd_auto(cfg, args) -> None:
    """
    Fa tutto in una volta:
      1. ricostruisce la coda dal calendario Excel
      2. riallinea il registro con quello che Facebook ha davvero in coda
      3. programma su Facebook Reel e caroselli fino a --days
      4. pubblica adesso cio' che e' scaduto (Storie comprese)
    """
    print("=== 1. Leggo il calendario ===")
    build_queue(cfg)

    if cfg.get("facebook", {}).get("enabled"):
        problema = verifica_token_fb(cfg)
        if problema:
            print(f"\n>>> TOKEN NON VALIDO: {problema}")
            print("    py ddp_publisher.py token --config config.json")
            return

        print("\n=== 2. Controllo cosa c'e' gia' su Facebook ===")
        cmd_sync(cfg)

        print(f"\n=== 3. Programmo i prossimi {args.days} giorni ===")
        cmd_fb_batch(cfg, args.days, args.dry_run)

    if cfg.get("instagram", {}).get("enabled"):
        print("\n=== 3b. Controllo cosa c'e' gia' su Instagram ===")
        cmd_ig_sync(cfg)

    print("\n=== 4. Pubblico cio' che e' scaduto adesso (Storie comprese) ===")
    cmd_run(cfg, args)

    print("\n=== FINITO ===")
    cmd_status(cfg)


def cmd_fb_batch(cfg, days: int, dry: bool) -> None:
    """Facebook accetta la programmazione nativa: carico in blocco il prossimo mese."""
    if not cfg.get("facebook", {}).get("enabled"):
        sys.exit("Facebook non abilitato nel config.")

    if not dry:
        problema = verifica_token_fb(cfg)
        if problema:
            print(f"\n>>> CONTROLLO TOKEN FALLITO:\n     {problema}")
            return
        print("Token della Pagina: verificato.\n")

    tasks = load_queue(cfg)
    state = load_state(cfg)
    fb = FacebookClient(cfg); fb.dry = dry
    now = datetime.now()

    # Facebook rifiuta le programmazioni oltre circa 30 giorni.
    # Taglio a 29 per stare larghi, invece di collezionare errori.
    MAX_GIORNI = 29
    limite_fb = now + timedelta(days=MAX_GIORNI)
    limit = now + timedelta(days=days)
    if limit > limite_fb:
        oltre = [t for t in tasks
                 if t.platform == "facebook" and t.key not in state["done"]
                 and t.fmt != "storia" and limite_fb < t.dt <= limit]
        limit = limite_fb
        if oltre:
            print(f"NOTA: Facebook non accetta programmazioni oltre ~{MAX_GIORNI} giorni.")
            print(f"      {len(oltre)} contenuti oltre il {limite_fb:%d/%m/%Y} li lascio")
            print(f"      per la prossima volta. Rilancia fra tre o quattro settimane")
            print(f"      e li caricherà da solo.\n")

    da_fare = [t for t in tasks
               if t.platform == "facebook"
               and t.key not in state["done"]
               and t.fmt != "storia"
               and now + timedelta(minutes=15) < t.dt <= limit]

    if not da_fare:
        print("Niente da programmare in questo periodo.")
        return

    print(f"{len(da_fare)} contenuti da programmare fino al {limit:%d/%m/%Y}.")
    print("Fra un caricamento e l'altro aspetto qualche secondo, per non")
    print("sovraccaricare il server che ospita i file.\n")

    ok = errori = 0
    for n, t in enumerate(da_fare, 1):
        urls = [media_url(cfg, m) for m in t.media]
        print(f"[{n}/{len(da_fare)}] {t.dt:%d/%m %H:%M}  {t.fmt:9} "
              f"{t.content_id} - {t.title}")
        try:
            if t.fmt == "reel":
                locale = Path(cfg["packs_root"]) / t.media[0]
                r = fb.reel(urls[0], t.caption, when=t.dt, file_locale=locale)
            else:
                r = fb.carousel(urls, t.caption, when=t.dt)
            info = f"fb schedulato {r.get('id') or r.get('post_id')}"
            print(f"    OK: {info}")
            ok += 1
            if not dry:
                mark(state, t, True, info)
                save_state(cfg, state)
        except Exception as e:
            testo = str(e)
            print(f"    ERRORE: {testo}")
            errori += 1
            if not dry:
                mark(state, t, False, testo)
                save_state(cfg, state)
            # Se il token e' scaduto, inutile insistere sugli altri 90 contenuti.
            if "scheduled publish time is invalid" in testo.lower():
                print("      (data troppo lontana per Facebook: la riprendo "
                      "al prossimo lancio)")
                state["failed"].pop(t.key, None)
                save_state(cfg, state)
                errori -= 1
                continue
            if any(s in testo.lower() for s in
                   ("access token", "session has expired", "oauthexception",
                    "validating access token", "code\": 190", "(#190")):
                print("\n>>> IL TOKEN E' SCADUTO. Mi fermo qui.")
                print(f"    Fatti finora: {ok}. Gli altri sono ancora da fare.")
                print("\n    Rigenera il token:")
                print("      py ddp_publisher.py token --config config.json")
                print("    poi rilancia lo stesso comando: riprende da dove era.")
                break
        if not dry and n < len(da_fare):
            time.sleep(12 if t.fmt == "reel" else 6)

    print(f"\nProgrammati: {ok}   Falliti: {errori}")
    if errori:
        print("Per i falliti rilancia lo stesso comando: riprende solo quelli.")


# ---------------------------------------------------------------------------


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("command",
                    choices=["build", "check", "plan", "run", "test",
                             "export", "fb-batch", "status",
                             "token", "token-check", "fb-list",
                             "sync", "auto", "token-instagram", "ig-sync",
                             "ig-refresh", "reset", "token-tiktok", "tiktok-batch"])
    ap.add_argument("--config", default="config.json")
    ap.add_argument("--days", type=int, default=7,
                    help="quanti giorni avanti programmare (max 180)")
    ap.add_argument("--window", type=int, default=30,
                    help="minuti di tolleranza a ritroso per i task scaduti")
    ap.add_argument("--dry-run", action="store_true",
                    help="stampa le chiamate senza inviare niente")

    g = ap.add_argument_group("filtri")
    g.add_argument("--only", metavar="PIATTAFORME",
                   help="instagram, facebook, tiktok (anche piu' d'una separata da virgola)")
    g.add_argument("--book", metavar="LIBRO",
                   help="filtra per libro, anche parziale: 'Notti'")
    g.add_argument("--content", metavar="ID",
                   help="TRAILER, CH01, CH02... separati da virgola")
    g.add_argument("--fmt", metavar="FORMATO",
                   help="reel, storia, carosello")
    g.add_argument("--task", metavar="ID",
                   help="numero preciso del task: T001")
    g.add_argument("--limit", type=int, default=0,
                   help="carica al massimo N video (0 = tutti)")
    g.add_argument("--yes", action="store_true",
                   help="salta la richiesta di conferma in 'test'")

    args = ap.parse_args()
    cfg = load_config(Path(args.config))

    if args.command == "build":
        build_queue(cfg)
    elif args.command == "check":
        sys.exit(cmd_check(cfg))
    elif args.command == "plan":
        cmd_plan(cfg, args)
    elif args.command == "run":
        cmd_run(cfg, args)
    elif args.command == "test":
        cmd_test(cfg, args)
    elif args.command == "export":
        cmd_export(cfg, args)
    elif args.command == "fb-batch":
        cmd_fb_batch(cfg, args.days, args.dry_run)
    elif args.command == "status":
        cmd_status(cfg)
    elif args.command == "token":
        cmd_token(cfg, args)
    elif args.command == "token-check":
        sys.exit(cmd_token_check(cfg))
    elif args.command == "fb-list":
        cmd_fb_list(cfg)
    elif args.command == "sync":
        cmd_sync(cfg)
    elif args.command == "auto":
        cmd_auto(cfg, args)
    elif args.command == "token-instagram":
        cmd_token_instagram(cfg, args)
    elif args.command == "ig-sync":
        cmd_ig_sync(cfg)
    elif args.command == "ig-refresh":
        cmd_ig_refresh(cfg)
    elif args.command == "reset":
        cmd_reset(cfg, args)
    elif args.command == "token-tiktok":
        cmd_token_tiktok(cfg, args)
    elif args.command == "tiktok-batch":
        cmd_tiktok_batch(cfg, args)


if __name__ == "__main__":
    main()